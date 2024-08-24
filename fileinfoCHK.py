#!python3
# -*- coding: utf_8 -*-

##import
#StandardLibrarys
import os
import sys
import math
#import time
import datetime
import configparser
#from collections import namedtuple
import logging

#ExternalLibrarys
#import numpy as np
#import tqdm
import openpyxl

##/import

#GlobalVariable
Dict_File_Info = {}

#/GlobalVariable

def readConfig():
    try:
        path = "config.ini"
        is_file = os.path.isfile(path)
        if not is_file:
            raise FileNotFoundError

        config = configparser.ConfigParser()
        config.read(path, encoding='utf-8')
        return config
    except (configparser.Error, FileNotFoundError) as e:
        pass

def progress_bar(cnt, files):
    progress = cnt / files
    percent = progress * 100

    bar_length = 50
    elapsed_len = int(bar_length * progress)
    bar = '#' * elapsed_len + '-' * (bar_length - elapsed_len)
    sys.stdout.write("\033[2K\033[G")
    sys.stdout.flush()
    print("\r"f'[{bar}] {cnt} / {files} ({percent}%)', end="")

def getMetadata(path):
    
    if os.stat(path).st_size >= 1000 and os.stat(path).st_size < 1000000:
        unit = 1000
        str_unit = "KByte"
    elif os.stat(path).st_size >= 1000000 and os.stat(path).st_size < 1000000000:
        unit = 1000000
        str_unit = "MByte"
    elif os.stat(path).st_size >= 1000000000:
        unit = 1000000000
        str_unit = "GByte"
    else:
        unit = 1
        str_unit = "Byte"
    
    metadata = {'ATIME': datetime.datetime.fromtimestamp(os.stat(path).st_atime),
                'MTIME': datetime.datetime.fromtimestamp(os.stat(path).st_mtime), 
                'CTIME': datetime.datetime.fromtimestamp(os.stat(path).st_ctime),
                'SIZE' : str(f'{float(os.stat(path).st_size / unit)} {str_unit}')}
    
    return metadata

def list_files(directory):
    global Dict_File_Info
    
    file_cnt = 0
    print(f'Directory Search >> {directory}')
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_cnt += 1

    cnt_main = 0
    for root, dirs, files in os.walk(directory):
        for file in files:
            filePath = str(os.path.join(root, file))
            Dict_File_Info[filePath] = getMetadata(filePath)
            cnt_main += 1
        
        progress_bar(cnt_main, file_cnt)    #Progress Check

def getTimeStamp():
    t_delta = datetime.timedelta(hours=9)
    JST = datetime.timezone(t_delta, 'JST')
    now = datetime.datetime.now(JST)
    return now

def outFile_xlsx():
    global Dict_File_Info
    rowLimit = 50000
    progress_bar_cnt = 0

    makecnt = 1
    dataMaxCnt = len(Dict_File_Info)
    if dataMaxCnt > rowLimit: makecnt = math.ceil(dataMaxCnt / rowLimit)

    pathList = [key for key in Dict_File_Info.keys()]

    wb = openpyxl.Workbook()
    sheet_index = 1

    print("Creating the output file")

    header = ["FilePath", "atime", "mtime", "ctime", "size"]

    for n in range(makecnt):
        ws = wb.create_sheet(f'Sheet{sheet_index}')
        sheet_index += 1

        #Header set
        for i, h in enumerate(header, 2):
            ws.cell(2, i).value = h
        
        #Data set
        dataLimit = 0
        for row_num, key in enumerate(pathList[(rowLimit * n):], 3):
            ws.cell(row_num, 2).value = key
            #ws.cell(i, 2).hyperlink = key  #240825:ハイパーリンクつけるとファイル開く際に無効化される
            ws.cell(row_num, 3).value = Dict_File_Info[key]['ATIME']
            ws.cell(row_num, 4).value = Dict_File_Info[key]['MTIME']
            ws.cell(row_num, 5).value = Dict_File_Info[key]['CTIME']
            ws.cell(row_num, 6).value = Dict_File_Info[key]['SIZE']

            progress_bar_cnt += 1
            progress_bar(progress_bar_cnt, dataMaxCnt)  #Progress Check

            if dataLimit >= rowLimit: break     #rowLimitで1シート区切る
            dataLimit += 1
        
    ws = wb.remove(wb['Sheet'])     #Delete unnecessary sheets
    print("\nnow saving...")
    now = getTimeStamp()            #timestamp
    wb.save(f'{now:%Y%m%d%H%M%S}.xlsx')
    print("File output successful")

def main():
    try:
        global Dict_File_Info

        configData = readConfig()
        if configData != None:
            for directory in configData['DIR_PATH'].values():
                list_files(directory)
                print("\n")
            else:
                outFile_xlsx()
                print('\nProcess complete!')
                input("Press Enter to continue...")
        else:
            raise FileNotFoundError('FileNotFoundError: Error reading config.ini')

    except Exception as e:
        print(e)
        input("Press Enter to continue...")

if __name__ == '__main__':
    main()
    